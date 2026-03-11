from rest_framework import viewsets, status, filters, permissions
from rest_framework.exceptions import ValidationError
from rest_framework.decorators import action
from rest_framework.response import Response
from django_filters.rest_framework import DjangoFilterBackend
from django.db import transaction
from django.db.models import Count, Prefetch, F, Subquery, OuterRef
from django.db.models.functions import TruncDate
from django.utils import timezone
from django.shortcuts import get_object_or_404
from django.conf import settings
from django.http import HttpResponse
from datetime import datetime, timedelta, time
import pytz
from openpyxl import Workbook

from .models import SalesAssignment, SalesPullRequest
from apps.calls.models import CallLog
from .serializers import (
    SalesAssignmentSerializer,
    SalesPullRequestSerializer,
    AssignmentHistorySummarySerializer,
    AssignmentHistoryDetailSerializer
)
from .services import assign_leads_to_agent, get_recycle_candidates
from .tasks import task_run_auto_assign
from apps.agents.models import Agent

class SalesAssignmentViewSet(viewsets.ModelViewSet):
    serializer_class = SalesAssignmentSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    http_method_names = ["get", "post", "patch", "head", "options"]
    
    # 필터링: 1차/2차 구분, 상태별 조회
    filterset_fields = ['stage', 'status', 'sentiment', 'agent']
    # 검색: 고객 이름, 전화번호로 검색
    search_fields = ['customer__name', 'customer__phone', 'memo']
    ordering_fields = ['updated_at', 'assigned_at']
    ordering = ['-updated_at'] # 최근 활동순 정렬

    def _parse_date(self, value):
        if not value:
            return None
        try:
            return datetime.strptime(value, "%Y-%m-%d").date()
        except ValueError:
            raise ValidationError({"detail": "날짜 형식이 올바르지 않습니다. (YYYY-MM-DD)"})

    def _get_date_range(self, request):
        start = self._parse_date(request.query_params.get('start_date'))
        end = self._parse_date(request.query_params.get('end_date'))

        today_kst = timezone.localtime().date()
        if end is None:
            end = today_kst
        if start is None:
            start = end - timedelta(days=6)

        if start > end:
            raise ValidationError({"detail": "start_date는 end_date보다 이후일 수 없습니다."})

        return start, end

    def _get_kst_range(self, start_date, end_date):
        kst = pytz.timezone('Asia/Seoul')
        start_dt = kst.localize(datetime.combine(start_date, time.min))
        end_dt = kst.localize(datetime.combine(end_date, time.max))

        if settings.USE_TZ:
            start_dt = start_dt.astimezone(pytz.UTC)
            end_dt = end_dt.astimezone(pytz.UTC)

        return start_dt, end_dt, kst

    def _get_history_queryset(self, request):
        stage = request.query_params.get('stage') or SalesAssignment.Stage.FIRST
        start_date, end_date = self._get_date_range(request)
        start_dt, end_dt, kst = self._get_kst_range(start_date, end_date)

        qs = SalesAssignment.objects.select_related(
            'agent', 'agent__user', 'customer'
        ).filter(
            agent__isnull=False,
            stage=stage,
            assigned_at__range=(start_dt, end_dt)
        )

        agent_id = request.query_params.get('agent_id')
        user = request.user

        if getattr(user, 'is_superuser', False):
            if agent_id:
                qs = qs.filter(agent_id=agent_id)
            return qs, start_date, end_date, kst

        if not hasattr(user, 'agent_profile'):
            return qs.none(), start_date, end_date, kst

        agent = user.agent_profile
        if agent.role in ['ADMIN', 'MANAGER']:
            if agent_id:
                qs = qs.filter(agent_id=agent_id)
        else:
            qs = qs.filter(agent=agent)

        return qs, start_date, end_date, kst

    def get_queryset(self):
        user = self.request.user
        last_memo_subquery = Subquery(
            CallLog.objects.filter(assignment=OuterRef('pk'))
            .order_by('-call_start').values('memo')[:1]
        )
        base_qs = SalesAssignment.objects.select_related('customer', 'agent', 'agent__user').annotate(
            call_count=Count('call_logs'),
            last_memo=last_memo_subquery
        ).prefetch_related(
            Prefetch(
                'child_assignments',
                queryset=SalesAssignment.objects.filter(stage='2ND').select_related('agent', 'agent__user'),
                to_attr='secondary_assignments'
            )
        )

        secondary_status = self.request.query_params.get('secondary_status')
        secondary_agent = self.request.query_params.get('secondary_agent')
        if secondary_status:
            base_qs = base_qs.filter(
                child_assignments__stage='2ND',
                child_assignments__status=secondary_status
            )
        if secondary_agent:
            base_qs = base_qs.filter(
                child_assignments__stage='2ND',
                child_assignments__agent_id=secondary_agent
            )
        if secondary_status or secondary_agent:
            base_qs = base_qs.distinct()
        if getattr(user, 'is_superuser', False):
            qs = base_qs
        else:
            if not hasattr(user, 'agent_profile'):
                return base_qs.none() # 상담원 아니면 빈 깡통

            agent = user.agent_profile
            
            # 관리자라면 전체 조회, 상담원이라면 '내 담당'만 조회
            if agent.role in ['ADMIN', 'MANAGER']:
                qs = base_qs
            else:
                qs = base_qs.filter(agent=agent)

        # 금일 배정된 건만 보기 (KST 기준)
        assigned_today = self.request.query_params.get('assigned_today')
        if assigned_today in ['1', 'true', 'True', 'yes', 'Y']:
            today_kst = timezone.localtime().date()
            kst = timezone.get_current_timezone()
            start_dt = timezone.make_aware(datetime.combine(today_kst, time.min), kst)
            end_dt = timezone.make_aware(datetime.combine(today_kst, time.max), kst)
            qs = qs.filter(assigned_at__range=(start_dt, end_dt))

        return qs

    @action(detail=False, methods=['get'], url_path='today-stats')
    def today_stats(self, request):
        user = request.user
        qs = SalesAssignment.objects.filter(stage=SalesAssignment.Stage.FIRST)

        if not getattr(user, 'is_superuser', False):
            if not hasattr(user, 'agent_profile'):
                return Response({'total': 0, 'completed': 0, 'remaining': 0})
            agent = user.agent_profile
            if agent.role not in ['ADMIN', 'MANAGER']:
                qs = qs.filter(agent=agent)

        today_kst = timezone.localtime().date()
        kst = timezone.get_current_timezone()
        start_dt = timezone.make_aware(datetime.combine(today_kst, time.min), kst)
        end_dt = timezone.make_aware(datetime.combine(today_kst, time.max), kst)
        qs = qs.filter(assigned_at__range=(start_dt, end_dt))

        total = qs.count()
        completed = qs.filter(status__in=['REJECT', 'INVALID', 'SUCCESS', 'BUY', 'HOLD']).count()

        return Response({
            'total': total,
            'completed': completed,
            'remaining': total - completed,
        })

    @action(detail=False, methods=['get'], url_path='recycle-candidates')
    def recycle_candidates(self, request):
        user = request.user
        if not getattr(user, 'is_superuser', False):
            if not hasattr(user, 'agent_profile'):
                return Response({"detail": "권한이 없습니다."}, status=status.HTTP_403_FORBIDDEN)
            agent = user.agent_profile
            if agent.role not in ['ADMIN', 'MANAGER']:
                return Response({"detail": "권한이 없습니다."}, status=status.HTTP_403_FORBIDDEN)

        qs = get_recycle_candidates(limit=None).filter(stage=SalesAssignment.Stage.FIRST)
        qs = qs.select_related('customer', 'agent', 'agent__user').annotate(
            call_count=Count('call_logs')
        ).prefetch_related(
            Prefetch(
                'child_assignments',
                queryset=SalesAssignment.objects.filter(stage='2ND').select_related('agent', 'agent__user'),
                to_attr='secondary_assignments'
            )
        )

        qs = self.filter_queryset(qs)
        page = self.paginate_queryset(qs)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)

        serializer = self.get_serializer(qs, many=True)
        return Response(serializer.data)

    def _export_assignments(self, qs, filename_prefix):
        wb = Workbook()
        ws = wb.active
        ws.title = "Assignments"

        ws.append([
            "배정ID",
            "단계",
            "단계표시",
            "상태",
            "상태표시",
            "배정일시",
            "수정일시",
            "1차 담당자ID",
            "1차 담당자명",
            "1차 담당자코드",
            "2차 배정ID",
            "2차 상태",
            "2차 상태표시",
            "2차 담당자ID",
            "2차 담당자명",
            "2차 담당자코드",
            "고객ID",
            "고객명",
            "전화번호",
            "나이",
            "성별",
            "지역",
            "분야1",
            "분야2",
            "분야3",
            "지역1",
            "지역2",
            "재활용횟수",
            "감도",
            "관심아이템",
            "메모",
            "통화횟수",
        ])

        for assignment in qs:
            secondary = None
            if hasattr(assignment, 'secondary_assignments'):
                secondary = assignment.secondary_assignments[0] if assignment.secondary_assignments else None
            else:
                secondary = assignment.child_assignments.filter(stage='2ND').select_related('agent', 'agent__user').first()

            ws.append([
                assignment.id,
                assignment.stage,
                assignment.get_stage_display(),
                assignment.status,
                assignment.get_status_display(),
                timezone.localtime(assignment.assigned_at).isoformat() if assignment.assigned_at else "",
                timezone.localtime(assignment.updated_at).isoformat() if assignment.updated_at else "",
                str(assignment.agent.agent_id) if assignment.agent else "",
                assignment.agent.user.name if assignment.agent and assignment.agent.user else "",
                assignment.agent.code if assignment.agent else "",
                secondary.id if secondary else "",
                secondary.status if secondary else "",
                secondary.get_status_display() if secondary else "",
                str(secondary.agent.agent_id) if secondary and secondary.agent else "",
                secondary.agent.user.name if secondary and secondary.agent and secondary.agent.user else "",
                secondary.agent.code if secondary and secondary.agent else "",
                assignment.customer.id if assignment.customer else "",
                assignment.customer.name if assignment.customer else "",
                assignment.customer.phone if assignment.customer else "",
                assignment.customer.age if assignment.customer else "",
                assignment.customer.gender if assignment.customer else "",
                assignment.customer.region if assignment.customer else "",
                assignment.customer.category_1 if assignment.customer else "",
                assignment.customer.category_2 if assignment.customer else "",
                assignment.customer.category_3 if assignment.customer else "",
                assignment.customer.region_1 if assignment.customer else "",
                assignment.customer.region_2 if assignment.customer else "",
                assignment.customer.recycle_count if assignment.customer else "",
                assignment.sentiment or "",
                assignment.item_interested or "",
                assignment.memo or "",
                assignment.call_count if hasattr(assignment, 'call_count') else 0,
            ])

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"{filename_prefix}_{timezone.localtime(timezone.now()).date().isoformat()}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response

    @action(detail=False, methods=['get'], url_path='success/export')
    def success_export(self, request):
        qs = self.get_queryset().filter(
            stage=SalesAssignment.Stage.FIRST,
            status=SalesAssignment.Status.SUCCESS
        )
        qs = self.filter_queryset(qs)
        return self._export_assignments(qs, "success_db")

    @action(detail=False, methods=['get'], url_path='recycle/export')
    def recycle_export(self, request):
        qs = get_recycle_candidates(limit=None).filter(stage=SalesAssignment.Stage.FIRST)
        qs = qs.select_related('customer', 'agent', 'agent__user').annotate(
            call_count=Count('call_logs')
        ).prefetch_related(
            Prefetch(
                'child_assignments',
                queryset=SalesAssignment.objects.filter(stage='2ND').select_related('agent', 'agent__user'),
                to_attr='secondary_assignments'
            )
        )
        qs = self.filter_queryset(qs)
        return self._export_assignments(qs, "recycle_db")

    @action(detail=False, methods=['post'])
    def pull(self, request):
        """
        🎯 [땡겨오기] 버튼 클릭 시 호출
        내게 할당된 1차 DB가 부족하면 추가로 가져옴
        """
        agent = request.user.agent_profile
        count = int(request.data.get('count', 10)) # 기본 10개
        
        # 서비스 로직 호출 (3단계에서 만듦)
        assigned_count = assign_leads_to_agent(agent, count=count)
        
        return Response({
            "message": f"{assigned_count}건의 DB를 새로 배정받았습니다.",
            "count": assigned_count
        }, status=status.HTTP_200_OK)

    @action(detail=False, methods=['post'], url_path='bulk-assign')
    def bulk_assign(self, request):
        ids = request.data.get('ids', [])
        agent_id = request.data.get('agent_id')

        if not ids:
            return Response({"detail": "배정할 데이터를 선택해주세요."}, status=status.HTTP_400_BAD_REQUEST)
        if not agent_id:
            return Response({"detail": "배정할 상담원을 선택해주세요."}, status=status.HTTP_400_BAD_REQUEST)

        agent = get_object_or_404(Agent, pk=agent_id)
        with transaction.atomic():
            updated_count = SalesAssignment.objects.filter(id__in=ids).update(
                agent=agent,
                status='ASSIGNED',
                assigned_at=timezone.now(),
                updated_at=timezone.now()
            )

        return Response({
            "message": f"성공적으로 {updated_count}건을 {agent.user.name}님에게 배정했습니다.",
            "updated_count": updated_count
        })

    @action(detail=False, methods=['post'], url_path='bulk-unassign')
    def bulk_unassign(self, request):
        ids = request.data.get('ids', [])
        if not ids:
            return Response({"detail": "선택된 데이터가 없습니다."}, status=status.HTTP_400_BAD_REQUEST)

        with transaction.atomic():
            updated_count = SalesAssignment.objects.filter(id__in=ids).update(
                agent=None,
                status='NEW',
                updated_at=timezone.now()
            )

        return Response({
            "message": f"✅ {updated_count}건의 배정이 취소되었습니다.",
            "updated_count": updated_count
        })

    @action(detail=False, methods=['post'], url_path='bulk-delete')
    def bulk_delete(self, request):
        ids = request.data.get('ids', [])
        if not ids:
            return Response({"detail": "삭제할 데이터를 선택해주세요."}, status=status.HTTP_400_BAD_REQUEST)

        with transaction.atomic():
            child_deleted = SalesAssignment.objects.filter(parent_assignment_id__in=ids).delete()[0]
            parent_deleted = SalesAssignment.objects.filter(id__in=ids).delete()[0]

        return Response({
            "message": f"✅ {parent_deleted}건 삭제, 2차 {child_deleted}건 삭제 완료",
            "deleted_count": parent_deleted,
            "deleted_secondary_count": child_deleted
        })

    @action(detail=False, methods=['post'], url_path='run-daily-assign')
    def run_daily_assign(self, request):
        trigger_user = request.user.name if hasattr(request.user, 'name') else request.user.username
        task_run_auto_assign.delay(triggered_by=trigger_user)

        return Response({
            'message': '자동 배정 작업이 시작되었습니다.',
            'info': '결과는 [자동 배정 이력] 메뉴에서 잠시 후 확인 가능합니다.'
        }, status=202)

    @action(detail=True, methods=['post'], url_path='assign-secondary')
    def assign_secondary(self, request, pk=None):
        assignment = self.get_object()
        agent_id = request.data.get('agent_id')

        if assignment.stage != SalesAssignment.Stage.FIRST:
            return Response({"detail": "1차 배정 건만 2차 배정이 가능합니다."}, status=status.HTTP_400_BAD_REQUEST)
        if assignment.status != SalesAssignment.Status.SUCCESS:
            return Response({"detail": "SUCCESS 상태인 건만 2차 배정이 가능합니다."}, status=status.HTTP_400_BAD_REQUEST)
        if assignment.child_assignments.filter(stage='2ND').exists():
            return Response({"detail": "이미 2차 배정이 존재합니다."}, status=status.HTTP_400_BAD_REQUEST)
        if not agent_id:
            return Response({"detail": "2차 담당자를 선택해주세요."}, status=status.HTTP_400_BAD_REQUEST)

        agent = get_object_or_404(Agent, pk=agent_id)

        secondary = SalesAssignment.objects.create(
            customer=assignment.customer,
            agent=agent,
            parent_assignment=assignment,
            stage=SalesAssignment.Stage.SECOND,
            status=SalesAssignment.Status.ASSIGNED,
        )

        serializer = self.get_serializer(secondary)
        return Response(serializer.data, status=status.HTTP_201_CREATED)

    @action(detail=False, methods=['get'], url_path='assignment-history/summary')
    def assignment_history_summary(self, request):
        qs, _, _, kst = self._get_history_queryset(request)

        summary = qs.annotate(
            date=TruncDate('assigned_at', tzinfo=kst),
            agent_name=F('agent__user__name')
        ).values(
            'date', 'agent_id', 'agent_name'
        ).annotate(
            assigned_count=Count('id')
        ).order_by('date', 'agent_name')

        serializer = AssignmentHistorySummarySerializer(summary, many=True)
        return Response({"items": serializer.data})

    @action(detail=False, methods=['get'], url_path='assignment-history/detail')
    def assignment_history_detail(self, request):
        qs, _, _, _ = self._get_history_queryset(request)

        detail = qs.annotate(
            agent_name=F('agent__user__name'),
            customer_name=F('customer__name')
        ).values(
            assignment_id=F('id'),
            assigned_at=F('assigned_at'),
            agent_id=F('agent_id'),
            agent_name=F('agent_name'),
            customer_id=F('customer_id'),
            customer_name=F('customer_name'),
            status=F('status')
        ).order_by('assigned_at')

        serializer = AssignmentHistoryDetailSerializer(detail, many=True)
        return Response({"items": serializer.data})

    @action(detail=False, methods=['get'], url_path='assignment-history/export')
    def assignment_history_export(self, request):
        qs, start_date, end_date, kst = self._get_history_queryset(request)

        summary = qs.annotate(
            date=TruncDate('assigned_at', tzinfo=kst),
            agent_name=F('agent__user__name')
        ).values(
            'date', 'agent_id', 'agent_name'
        ).annotate(
            assigned_count=Count('id')
        )

        summary_map = {}
        for row in summary:
            key = (str(row['agent_id']), row['date'].isoformat())
            summary_map[key] = row['assigned_count']

        if getattr(request.user, 'is_superuser', False) or (
            hasattr(request.user, 'agent_profile') and request.user.agent_profile.role in ['ADMIN', 'MANAGER']
        ):
            agent_id = request.query_params.get('agent_id')
            agents_qs = Agent.objects.select_related('user').all()
            if agent_id:
                agents_qs = agents_qs.filter(agent_id=agent_id)
        else:
            agents_qs = Agent.objects.select_related('user').filter(user=request.user)

        agents = list(agents_qs.order_by('created_at'))

        date_cursor = start_date
        dates = []
        while date_cursor <= end_date:
            dates.append(date_cursor)
            date_cursor += timedelta(days=1)

        wb = Workbook()
        ws_pivot = wb.active
        ws_pivot.title = "Pivot"

        header = ["상담원"] + [d.isoformat() for d in dates]
        ws_pivot.append(header)

        for agent in agents:
            row = [agent.user.name]
            for d in dates:
                key = (str(agent.agent_id), d.isoformat())
                row.append(summary_map.get(key, 0))
            ws_pivot.append(row)

        ws_detail = wb.create_sheet("Detail")
        ws_detail.append([
            "배정ID", "배정일시", "상담원ID", "상담원명", "고객ID", "고객명", "상태"
        ])

        detail = qs.annotate(
            agent_name=F('agent__user__name'),
            customer_name=F('customer__name')
        ).values(
            assignment_id=F('id'),
            assigned_at=F('assigned_at'),
            agent_id=F('agent_id'),
            agent_name=F('agent_name'),
            customer_id=F('customer_id'),
            customer_name=F('customer_name'),
            status=F('status')
        ).order_by('assigned_at')

        for row in detail:
            ws_detail.append([
                row['assignment_id'],
                timezone.localtime(row['assigned_at']).isoformat(),
                str(row['agent_id']),
                row['agent_name'],
                row['customer_id'],
                row['customer_name'],
                row['status']
            ])

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"assignment_history_{start_date.isoformat()}_{end_date.isoformat()}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response


# [2] 땡겨오기 신청/승인 ViewSet
class SalesPullRequestViewSet(viewsets.ModelViewSet):
    serializer_class = SalesPullRequestSerializer
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    http_method_names = ["get", "post", "head", "options"]
    filterset_fields = ['status', 'agent']
    ordering_fields = ['created_at', 'processed_at']
    ordering = ['-created_at']

    def _parse_date(self, value):
        if not value:
            return None
        try:
            return datetime.strptime(value, "%Y-%m-%d").date()
        except ValueError:
            raise ValidationError({"detail": "날짜 형식이 올바르지 않습니다. (YYYY-MM-DD)"})

    def _get_kst_range(self, start_date, end_date):
        kst = pytz.timezone('Asia/Seoul')
        start_dt = kst.localize(datetime.combine(start_date, time.min))
        end_dt = kst.localize(datetime.combine(end_date, time.max))

        if settings.USE_TZ:
            start_dt = start_dt.astimezone(pytz.UTC)
            end_dt = end_dt.astimezone(pytz.UTC)

        return start_dt, end_dt

    def get_queryset(self):
        user = self.request.user
        base_qs = SalesPullRequest.objects.select_related(
            'agent', 'agent__user', 'processed_by', 'processed_by__user'
        )

        start_date = self._parse_date(self.request.query_params.get('start_date'))
        end_date = self._parse_date(self.request.query_params.get('end_date'))
        if start_date or end_date:
            if start_date is None:
                start_date = end_date
            if end_date is None:
                end_date = start_date
            if start_date > end_date:
                raise ValidationError({"detail": "start_date는 end_date보다 이후일 수 없습니다."})
            start_dt, end_dt = self._get_kst_range(start_date, end_date)
            base_qs = base_qs.filter(created_at__range=(start_dt, end_dt))

        if getattr(user, 'is_superuser', False):
            return base_qs
        if not hasattr(user, 'agent_profile'):
            return base_qs.none()

        agent = user.agent_profile
        if agent.role in ['ADMIN', 'MANAGER']:
            return base_qs
        return base_qs.filter(agent=agent)

    def perform_create(self, serializer):
        if not hasattr(self.request.user, 'agent_profile'):
            raise ValidationError({"detail": "상담원 프로필이 필요합니다."})
        serializer.save(agent=self.request.user.agent_profile)

    def _ensure_admin(self, request):
        if getattr(request.user, 'is_superuser', False):
            return True
        if not hasattr(request.user, 'agent_profile'):
            return False
        return request.user.agent_profile.role in ['ADMIN', 'MANAGER']

    @action(detail=True, methods=['post'])
    def approve(self, request, pk=None):
        if not self._ensure_admin(request):
            return Response({"detail": "권한이 없습니다."}, status=status.HTTP_403_FORBIDDEN)

        pull_request = self.get_object()
        if pull_request.status != SalesPullRequest.Status.PENDING:
            return Response({"detail": "이미 처리된 요청입니다."}, status=status.HTTP_400_BAD_REQUEST)

        assigned_count = assign_leads_to_agent(pull_request.agent, count=pull_request.requested_count)
        pull_request.status = SalesPullRequest.Status.APPROVED
        pull_request.approved_count = assigned_count
        pull_request.processed_by = getattr(request.user, 'agent_profile', None)
        pull_request.processed_at = timezone.now()
        pull_request.save(update_fields=[
            'status', 'approved_count', 'processed_by', 'processed_at', 'updated_at'
        ])

        return Response({
            "message": f"{assigned_count}건을 배정했습니다.",
            "assigned_count": assigned_count,
            "request_id": pull_request.id
        }, status=status.HTTP_200_OK)

    @action(detail=True, methods=['post'])
    def reject(self, request, pk=None):
        if not self._ensure_admin(request):
            return Response({"detail": "권한이 없습니다."}, status=status.HTTP_403_FORBIDDEN)

        pull_request = self.get_object()
        if pull_request.status != SalesPullRequest.Status.PENDING:
            return Response({"detail": "이미 처리된 요청입니다."}, status=status.HTTP_400_BAD_REQUEST)

        reject_reason = request.data.get('reason', '') or ''
        pull_request.status = SalesPullRequest.Status.REJECTED
        pull_request.reject_reason = reject_reason
        pull_request.processed_by = getattr(request.user, 'agent_profile', None)
        pull_request.processed_at = timezone.now()
        pull_request.save(update_fields=[
            'status', 'reject_reason', 'processed_by', 'processed_at', 'updated_at'
        ])

        return Response({
            "message": "요청을 거절했습니다.",
            "request_id": pull_request.id
        }, status=status.HTTP_200_OK)
